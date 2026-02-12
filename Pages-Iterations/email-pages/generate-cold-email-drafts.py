#!/usr/bin/env python3
"""
Generate personalized cold-email drafts from XLSX + HTML template.

Outputs:
- Rendered HTML files for quick review.
- RFC822 .eml files with inline images (CID) for Gmail API draft creation.
- Base64url payload files (`raw/*.b64url.txt`) compatible with Gmail API.
- Manifest CSV/JSON with mapping details and warnings.
"""

from __future__ import annotations

import argparse
import base64
import csv
import html
import json
import mimetypes
import os
import re
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from email.message import EmailMessage
from pathlib import Path
from typing import Any

import openpyxl


DEFAULT_APP_URL = "https://apps.apple.com/us/app/echochat/id6736381852"
DEFAULT_LOCATION = "Marina, San Francisco"
DEFAULT_SUBJECT_TEMPLATE = "Echo found 3 memory matches for you, {user_display_name}"
DEFAULT_FROM = "Echo@iditor.com"

PLACEHOLDER_RE = re.compile(r"\{\{\s*([a-zA-Z0-9_]+)\s*\}\}")
ASSET_RE = re.compile(r"^(\d+)_(.+)_(highlight|match1|match2|match3)\.(jpg|jpeg|png|webp)$", re.IGNORECASE)
RAW_HTML_KEYS = {"match1_intro_html", "match2_intro_html", "match3_intro_html"}
MOJIBAKE_HINTS = ("Ã", "Â", "â€™", "â€œ", "â€", "â€“", "â€”", "â€¦")


@dataclass
class AssetGroup:
    index: int
    handle: str
    files: dict[str, Path]


def parse_args() -> argparse.Namespace:
    here = Path(__file__).resolve().parent
    parser = argparse.ArgumentParser(description="Generate Gmail-ready cold email drafts from XLSX.")
    parser.add_argument(
        "--xlsx",
        default=str(here / "cold_email_list.xlsx"),
        help="Path to source XLSX file.",
    )
    parser.add_argument(
        "--template",
        default=str(here / "email-template.html"),
        help="Path to HTML template.",
    )
    parser.add_argument(
        "--assets-dir",
        default=str(here / "generated" / "assets" / "images"),
        help="Directory containing generated image assets.",
    )
    parser.add_argument(
        "--output-dir",
        default=str(here / "generated" / "drafts" / "test-shot"),
        help="Output folder for html/eml/raw/manifest files.",
    )
    parser.add_argument("--sheet", default="", help="Optional Excel sheet name. Defaults to first sheet.")
    parser.add_argument("--start", type=int, default=1, help="1-based row index to start from (data rows only).")
    parser.add_argument("--limit", type=int, default=3, help="Number of rows to generate.")
    parser.add_argument("--location", default=DEFAULT_LOCATION, help="Fallback location placeholder.")
    parser.add_argument("--app-url", default=DEFAULT_APP_URL, help="App URL used in template.")
    parser.add_argument("--invite-code", default="ODLKV", help="Invite code shown in template.")
    parser.add_argument("--expiry-days", type=int, default=3, help="Invite expiry days.")
    parser.add_argument("--from-email", default=DEFAULT_FROM, help="From address used in MIME drafts.")
    parser.add_argument(
        "--subject-template",
        default=DEFAULT_SUBJECT_TEMPLATE,
        help="Python format string for subject, e.g. 'Hi {user_display_name}'.",
    )
    parser.add_argument(
        "--image-base-url",
        default="",
        help="Optional public base URL for local images in review HTML (if omitted, local relative paths are used).",
    )
    parser.add_argument(
        "--allow-index-fallback",
        action="store_true",
        help="Allow row-index image matching when username-based matching fails (disabled by default).",
    )
    return parser.parse_args()


def normalize_handle(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", "", (value or "").lower())


def slugify(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", "-", (value or "").lower()).strip("-")[:64]


def maybe_fix_mojibake(text: str) -> str:
    """Repair common UTF-8 bytes that were mis-decoded as CP1252."""
    value = str(text or "")
    fixed = value
    for _ in range(2):
        if not any(hint in fixed for hint in MOJIBAKE_HINTS):
            break
        try:
            repaired = fixed.encode("cp1252").decode("utf-8")
        except (UnicodeEncodeError, UnicodeDecodeError):
            break
        if repaired == fixed:
            break
        fixed = repaired
    return fixed


def deep_fix_strings(value: Any) -> Any:
    if isinstance(value, str):
        return maybe_fix_mojibake(value)
    if isinstance(value, list):
        return [deep_fix_strings(x) for x in value]
    if isinstance(value, dict):
        return {k: deep_fix_strings(v) for k, v in value.items()}
    return value


def parse_json_field(value: Any, warnings: list[str], field_name: str, row_num: int) -> dict[str, Any]:
    if isinstance(value, dict):
        return value
    raw = str(value or "").strip()
    if not raw:
        return {}
    try:
        parsed = json.loads(raw)
        parsed = deep_fix_strings(parsed)
        return parsed if isinstance(parsed, dict) else {}
    except json.JSONDecodeError:
        warnings.append(f"Row {row_num}: invalid JSON in '{field_name}'.")
        return {}


def format_date_mon_d(dt_value: date) -> str:
    return f"{dt_value.strftime('%b')} {dt_value.day}"


def parse_datetime(value: str) -> datetime | None:
    text = (value or "").strip()
    if not text:
        return None
    try:
        return datetime.fromisoformat(text.replace("Z", "+00:00"))
    except ValueError:
        return None


def format_memory_time(highlight: dict[str, Any]) -> str:
    dt_value = parse_datetime(str(highlight.get("time", "")))
    if not dt_value:
        return "Recently captured"
    return f"{dt_value.strftime('%b')} {dt_value.day}, {dt_value.year}"


def format_match_intro(value: str) -> str:
    raw = maybe_fix_mojibake(str(value or "")).strip()
    if not raw:
        return ""

    marks: list[str] = []

    def store_quote(match: re.Match[str]) -> str:
        marks.append(match.group(1).strip())
        return f"@@Q{len(marks) - 1}@@"

    raw = re.sub(r'"([^"]+)"', store_quote, raw)
    escaped = html.escape(raw)

    for idx, phrase in enumerate(marks):
        escaped = escaped.replace(
            f"@@Q{idx}@@",
            f'<span style="color:#111111;font-weight:700;">{html.escape(phrase)}</span>',
        )

    escaped = re.sub(
        r"((?:\b(?:one|two|three|four|five|six|seven|eight|nine|ten|\d+)\s+)?"
        r"(?:day|days|week|weeks|month|months|year|years)\s+(?:earlier|later|prior)),\s*",
        r"\1<br />",
        escaped,
        flags=re.IGNORECASE,
    )
    return escaped


def render_template(template: str, view: dict[str, str]) -> str:
    def repl(match: re.Match[str]) -> str:
        key = match.group(1)
        value = view.get(key, "")
        return str(value) if key in RAW_HTML_KEYS else html.escape(str(value))

    return PLACEHOLDER_RE.sub(repl, template)


def unresolved_placeholders(rendered_html: str) -> list[str]:
    return sorted(set(PLACEHOLDER_RE.findall(rendered_html)))


def build_asset_index(assets_dir: Path) -> tuple[dict[str, list[AssetGroup]], dict[int, list[AssetGroup]]]:
    groups: dict[tuple[int, str], AssetGroup] = {}
    if not assets_dir.exists():
        return {}, {}

    for file in assets_dir.iterdir():
        if not file.is_file():
            continue
        m = ASSET_RE.match(file.name)
        if not m:
            continue
        idx = int(m.group(1))
        handle = m.group(2)
        slot = m.group(3).lower()
        key = (idx, handle)
        if key not in groups:
            groups[key] = AssetGroup(index=idx, handle=handle, files={})
        groups[key].files[slot] = file

    by_norm: dict[str, list[AssetGroup]] = {}
    by_index: dict[int, list[AssetGroup]] = {}
    for group in groups.values():
        by_norm.setdefault(normalize_handle(group.handle), []).append(group)
        by_index.setdefault(group.index, []).append(group)
    return by_norm, by_index


def choose_asset_group(
    row_index: int,
    username: str,
    by_norm: dict[str, list[AssetGroup]],
    by_index: dict[int, list[AssetGroup]],
    allow_index_fallback: bool,
) -> tuple[AssetGroup | None, str]:
    norm = normalize_handle(username)

    direct = by_norm.get(norm, [])
    if direct:
        chosen = sorted(direct, key=lambda g: (-len(g.files), abs(g.index - row_index), g.handle))[0]
        return chosen, "username"

    fuzzy: list[AssetGroup] = []
    for key, groups in by_norm.items():
        if norm and (norm in key or key in norm):
            fuzzy.extend(groups)
    if fuzzy:
        chosen = sorted(fuzzy, key=lambda g: (-len(g.files), abs(g.index - row_index), g.handle))[0]
        return chosen, "fuzzy"

    if allow_index_fallback:
        by_row = by_index.get(row_index, [])
        if by_row:
            chosen = sorted(by_row, key=lambda g: (-len(g.files), g.handle))[0]
            return chosen, "index"

    return None, "none"


def read_xlsx_records(xlsx_path: Path, sheet_name: str) -> tuple[str, list[dict[str, Any]]]:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
    try:
        actual_sheet = sheet_name if sheet_name and sheet_name in wb.sheetnames else wb.sheetnames[0]
        ws = wb[actual_sheet]
        rows = list(ws.iter_rows(values_only=True))
    finally:
        wb.close()

    if len(rows) < 2:
        raise ValueError("XLSX must contain a header row and at least one data row.")

    headers = [str(h or "").strip() for h in rows[0]]
    records: list[dict[str, Any]] = []
    for row in rows[1:]:
        rec: dict[str, Any] = {}
        for i, key in enumerate(headers):
            rec[key] = row[i] if i < len(row) else ""
        records.append(rec)
    return actual_sheet, records


def view_for_record(
    rec: dict[str, Any],
    row_idx: int,
    group: AssetGroup | None,
    args: argparse.Namespace,
    warnings: list[str],
) -> tuple[dict[str, str], dict[str, Path]]:
    username = maybe_fix_mojibake(str(rec.get("username") or "")).strip()
    display_name = maybe_fix_mojibake(str(rec.get("name") or username or "there")).strip()
    email_to = str(rec.get("email") or "").strip()

    highlight = parse_json_field(rec.get("highlight"), warnings, "highlight", row_idx)
    top_1 = parse_json_field(rec.get("top_1"), warnings, "top_1", row_idx)
    top_2 = parse_json_field(rec.get("top_2"), warnings, "top_2", row_idx)
    top_3 = parse_json_field(rec.get("top_3"), warnings, "top_3", row_idx)
    promo = parse_json_field(rec.get("promo"), warnings, "promo", row_idx)

    promo_highlight = (
        maybe_fix_mojibake(str(rec.get("promo_highlight") or "")).strip()
        or maybe_fix_mojibake(str(highlight.get("description") or "")).strip()
    )
    if not promo_highlight:
        promo_highlight = "You left a memory that strongly resonated with others in Echo."

    expiry_date = date.today() + timedelta(days=max(0, args.expiry_days))

    local_images: dict[str, Path] = {}

    def image_for(slot: str, remote_fallback: str) -> str:
        local = group.files.get(slot) if group else None
        if local:
            local_images[slot] = local
            return f"@@LOCAL:{slot}@@"
        return str(remote_fallback or "")

    def top_values(top: dict[str, Any], slot: str, promo_key: str) -> dict[str, str]:
        owner = top.get("owner") if isinstance(top.get("owner"), dict) else {}
        handle = str(owner.get("twitterUsername") or "").strip().lstrip("@")
        if not handle and isinstance(top.get("twitter"), str):
            handle = str(top.get("twitter")).strip().lstrip("@")
        name = maybe_fix_mojibake(str(owner.get("name") or top.get("name") or "Match")).strip()
        avatar = str(owner.get("avatarUrl") or top.get("avatarUrl") or "").strip()
        intro = maybe_fix_mojibake(str(promo.get(promo_key) or top.get("excerpt") or "")).strip()
        return {
            "name": name,
            "twitter_url": f"https://x.com/{handle}" if handle else "https://x.com",
            "headshot": image_for(slot, avatar),
            "intro_html": format_match_intro(intro),
        }

    m1 = top_values(top_1, "match1", "match_1")
    m2 = top_values(top_2, "match2", "match_2")
    m3 = top_values(top_3, "match3", "match_3")

    view = {
        "user_display_name": display_name,
        "user_name": username,
        "user_email": email_to,
        "location": args.location,
        "promo_highlight": promo_highlight,
        "memory_time": format_memory_time(highlight),
        "memory_manga_image": image_for("highlight", str(highlight.get("url") or "")),
        "match1_name": m1["name"],
        "match1_twitter_url": m1["twitter_url"],
        "match1_headshot_png": m1["headshot"],
        "match1_intro_html": m1["intro_html"],
        "match2_name": m2["name"],
        "match2_twitter_url": m2["twitter_url"],
        "match2_headshot_png": m2["headshot"],
        "match2_intro_html": m2["intro_html"],
        "match3_name": m3["name"],
        "match3_twitter_url": m3["twitter_url"],
        "match3_headshot_png": m3["headshot"],
        "match3_intro_html": m3["intro_html"],
        "invite_code": args.invite_code,
        "expiry_days": str(args.expiry_days),
        "invite_expires_on": format_date_mon_d(expiry_date),
        "app_url": args.app_url,
    }
    return view, local_images


def local_or_public_image_ref(
    slot: str,
    local_images: dict[str, Path],
    html_dir: Path,
    image_base_url: str,
) -> str | None:
    local = local_images.get(slot)
    if not local:
        return None
    if image_base_url:
        return f"{image_base_url.rstrip('/')}/{local.name}"
    rel_path = os.path.relpath(local.resolve(), html_dir.resolve())
    return rel_path.replace("\\", "/")


def swap_local_image_tokens(
    view: dict[str, str],
    local_images: dict[str, Path],
    html_dir: Path,
    image_base_url: str,
    cid_mode: bool,
) -> tuple[dict[str, str], dict[str, str]]:
    replaced = dict(view)
    cid_map: dict[str, str] = {}

    mapping = {
        "memory_manga_image": "highlight",
        "match1_headshot_png": "match1",
        "match2_headshot_png": "match2",
        "match3_headshot_png": "match3",
    }
    for key, slot in mapping.items():
        token = f"@@LOCAL:{slot}@@"
        value = replaced.get(key, "")
        if value != token:
            continue

        local = local_images.get(slot)
        if not local:
            replaced[key] = ""
            continue

        if cid_mode:
            cid = f"{slot}-{local.stem}@echo.local"
            replaced[key] = f"cid:{cid}"
            cid_map[slot] = cid
            continue

        ref = local_or_public_image_ref(slot, local_images, html_dir, image_base_url)
        replaced[key] = ref or ""

    return replaced, cid_map


def ensure_dirs(base: Path) -> tuple[Path, Path, Path]:
    html_dir = base / "html"
    eml_dir = base / "eml"
    raw_dir = base / "raw"
    for d in (html_dir, eml_dir, raw_dir):
        d.mkdir(parents=True, exist_ok=True)
    return html_dir, eml_dir, raw_dir


def build_email_message(
    from_email: str,
    to_email: str,
    subject: str,
    html_body: str,
    local_images: dict[str, Path],
    cid_map: dict[str, str],
) -> EmailMessage:
    msg = EmailMessage()
    msg["From"] = from_email
    msg["To"] = to_email
    msg["Subject"] = subject
    msg.set_content("Open this email in HTML mode to view your personalized Echo invite.")
    msg.add_alternative(html_body, subtype="html")

    html_part = msg.get_payload()[-1]
    for slot, cid in cid_map.items():
        file = local_images.get(slot)
        if not file or not file.exists():
            continue
        mime_type = mimetypes.guess_type(file.name)[0] or "application/octet-stream"
        maintype, subtype = mime_type.split("/", 1)
        with file.open("rb") as f:
            data = f.read()
        html_part.add_related(
            data,
            maintype=maintype,
            subtype=subtype,
            cid=f"<{cid}>",
            filename=file.name,
            disposition="inline",
        )
    return msg


def main() -> None:
    args = parse_args()

    xlsx_path = Path(args.xlsx).resolve()
    template_path = Path(args.template).resolve()
    assets_dir = Path(args.assets_dir).resolve()
    output_dir = Path(args.output_dir).resolve()

    if not xlsx_path.exists():
        raise FileNotFoundError(f"XLSX not found: {xlsx_path}")
    if not template_path.exists():
        raise FileNotFoundError(f"Template not found: {template_path}")

    template = template_path.read_text(encoding="utf-8")
    sheet_name, records = read_xlsx_records(xlsx_path, args.sheet)
    by_norm, by_index = build_asset_index(assets_dir)
    html_dir, eml_dir, raw_dir = ensure_dirs(output_dir)

    start = max(1, args.start)
    selected = records[start - 1 :]
    if args.limit is not None and args.limit >= 0:
        selected = selected[: args.limit]

    manifest_rows: list[dict[str, str]] = []
    global_warnings: list[str] = []

    for pos, rec in enumerate(selected, start=start):
        username = str(rec.get("username") or f"user-{pos}").strip()
        email_to = str(rec.get("email") or "").strip()
        if not email_to:
            global_warnings.append(f"Row {pos}: missing email, skipped.")
            continue

        group, match_mode = choose_asset_group(
            pos,
            username,
            by_norm,
            by_index,
            allow_index_fallback=args.allow_index_fallback,
        )
        row_warnings: list[str] = []
        if not group:
            row_warnings.append(f"Row {pos}: no local image group matched '{username}'.")
        elif match_mode != "username":
            row_warnings.append(
                f"Row {pos}: image group matched via '{match_mode}' mode ({group.index}_{group.handle})."
            )

        view, local_images = view_for_record(rec, pos, group, args, row_warnings)

        subject = args.subject_template.format(**view)
        base = f"{pos:02d}-{slugify(username)}-{slugify(email_to.split('@')[0])}"
        if not base.strip("-"):
            base = f"{pos:02d}-recipient"

        review_view, _ = swap_local_image_tokens(
            view=view,
            local_images=local_images,
            html_dir=html_dir,
            image_base_url=args.image_base_url,
            cid_mode=False,
        )
        review_html = render_template(template, review_view)
        unresolved_review = unresolved_placeholders(review_html)
        if unresolved_review:
            row_warnings.append(f"Row {pos}: unresolved placeholders in review HTML: {', '.join(unresolved_review)}")
        html_path = html_dir / f"{base}.html"
        html_path.write_text(review_html, encoding="utf-8")

        gmail_view, cid_map = swap_local_image_tokens(
            view=view,
            local_images=local_images,
            html_dir=html_dir,
            image_base_url=args.image_base_url,
            cid_mode=True,
        )
        gmail_html = render_template(template, gmail_view)
        unresolved_gmail = unresolved_placeholders(gmail_html)
        if unresolved_gmail:
            row_warnings.append(f"Row {pos}: unresolved placeholders in Gmail HTML: {', '.join(unresolved_gmail)}")

        msg = build_email_message(
            from_email=args.from_email,
            to_email=email_to,
            subject=subject,
            html_body=gmail_html,
            local_images=local_images,
            cid_map=cid_map,
        )

        eml_path = eml_dir / f"{base}.eml"
        eml_bytes = msg.as_bytes()
        eml_path.write_bytes(eml_bytes)

        raw_b64 = base64.urlsafe_b64encode(eml_bytes).decode("ascii")
        raw_path = raw_dir / f"{base}.b64url.txt"
        raw_path.write_text(raw_b64, encoding="utf-8")

        missing_slots = [slot for slot in ("highlight", "match1", "match2", "match3") if slot not in local_images]
        manifest_rows.append(
            {
                "row_index": str(pos),
                "sheet": sheet_name,
                "username": username,
                "to_email": email_to,
                "subject": subject,
                "html_path": str(html_path),
                "eml_path": str(eml_path),
                "raw_b64url_path": str(raw_path),
                "asset_match_mode": match_mode,
                "asset_group": f"{group.index}_{group.handle}" if group else "",
                "missing_local_slots": "|".join(missing_slots),
                "warnings": " | ".join(row_warnings),
            }
        )
        global_warnings.extend(row_warnings)

    json_path = output_dir / "manifest.json"
    csv_path = output_dir / "manifest.csv"
    (output_dir / "warnings.txt").write_text("\n".join(global_warnings), encoding="utf-8")
    json_path.write_text(json.dumps(manifest_rows, indent=2), encoding="utf-8")

    fieldnames = [
        "row_index",
        "sheet",
        "username",
        "to_email",
        "subject",
        "html_path",
        "eml_path",
        "raw_b64url_path",
        "asset_match_mode",
        "asset_group",
        "missing_local_slots",
        "warnings",
    ]
    with csv_path.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(manifest_rows)

    print(f"Sheet: {sheet_name}")
    print(f"Rows processed: {len(manifest_rows)}")
    print(f"Output directory: {output_dir}")
    print(f"Manifest JSON: {json_path}")
    print(f"Manifest CSV: {csv_path}")
    if global_warnings:
        print(f"Warnings: {len(global_warnings)} (see {output_dir / 'warnings.txt'})")
    else:
        print("Warnings: 0")


if __name__ == "__main__":
    main()
